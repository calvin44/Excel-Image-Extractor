import os
import argparse
import shutil
import json
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from openpyxl.utils import get_column_letter

# pylint: disable=protected-access


def extract_images_from_excel(
    file_path: str,
    sheet_name: str,
    output_folder: str
) -> List[Dict[str, str]]:
    """
    Extract images from a specific sheet in an Excel (.xlsx) file and save them to disk.

    Clears the output folder before saving images.

    Args:
        file_path (str): Path to the Excel file (.xlsx).
        sheet_name (str): Name of the sheet to extract images from.
        output_folder (str): Absolute or relative folder path to save extracted images.

    Returns:
        List[Dict[str, str]]: A list of dicts with keys:
            - 'image_file': Absolute path to the saved image.
            - 'cell': Excel cell where the image is anchored (e.g., 'B3').
            - 'col': Column number (1-based).
            - 'row': Row number (1-based).

    Raises:
        ValueError: If the sheet name does not exist in the workbook.
    """
    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = wb[sheet_name]

    # Clear output folder if it exists, else create it
    if os.path.exists(output_folder):
        for filename in os.listdir(output_folder):
            file_path_inner = os.path.join(output_folder, filename)
            if os.path.isfile(file_path_inner) or os.path.islink(file_path_inner):
                os.unlink(file_path_inner)
            elif os.path.isdir(file_path_inner):
                shutil.rmtree(file_path_inner)
    else:
        os.makedirs(output_folder)

    results = []

    for idx, image in enumerate(ws._images, start=1):
        anchor = image.anchor

        if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
            col = anchor._from.col
            row = anchor._from.row
            col_num = col + 1
            row_num = row + 1
            col_letter = get_column_letter(col_num)
            cell = f"{col_letter}{row_num}"

            image_filename = f"image_{idx}_{cell}.png"
            image_path = os.path.abspath(
                os.path.join(output_folder, image_filename))

            with open(image_path, "wb") as f:
                f.write(image._data())

            results.append({
                "image_file": image_path,
                "cell": cell,
                "col": col_num,
                "row": row_num,
            })

    return results


def main():
    parser = argparse.ArgumentParser(
        description="Extract images from Excel sheet and save to folder.")
    parser.add_argument("excel_file", help="Path to the Excel (.xlsx) file")
    parser.add_argument("sheet_name", help="Sheet name to extract images from")
    parser.add_argument(
        "output_folder", help="Folder path to save extracted images")

    args = parser.parse_args()

    try:
        results = extract_images_from_excel(
            args.excel_file, args.sheet_name, args.output_folder)
    except Exception as e:  # pylint: disable=broad-except
        print(f"[X ERROR] {e}")
        exit(1)

    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main()
