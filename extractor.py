import os
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from utils import clear_folder, get_cell_position

# pylint: disable=protected-access


def extract_images_from_excel(
    file_path: str,
    sheet_name: str,
    output_folder: str
) -> List[Dict[str, str]]:
    """
    Extract images from a specific sheet in an Excel (.xlsx) file and save them to disk.

    Clears the output folder before saving images.

    Args:
        file_path (str): Path to the Excel file (.xlsx).
        sheet_name (str): Name of the sheet to extract images from.
        output_folder (str): Absolute or relative folder path to save extracted images.

    Returns:
        List[Dict[str, str]]: A list of dicts with keys:
            - 'image_file': Absolute path to the saved image.
            - 'cell': Excel cell where the image is anchored (e.g., 'B3').
            - 'col': Column number (1-based).
            - 'row': Row number (1-based).

    Raises:
        ValueError: If the sheet name does not exist in the workbook.
    """

    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = wb[sheet_name]
    clear_folder(output_folder)

    results = []

    for idx, image in enumerate(ws._images, start=1):
        anchor = image.anchor
        if not isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
            continue  # skip if anchor is not supported

        col, row, cell = get_cell_position(anchor)

        image_filename = f"image_{idx}_{cell}.png"
        image_path = os.path.abspath(
            os.path.join(output_folder, image_filename))

        with open(image_path, "wb") as f:
            f.write(image._data())

        results.append({
            "image_file": image_path,
            "cell": cell,
            "col": col,
            "row": row,
        })

    return results
